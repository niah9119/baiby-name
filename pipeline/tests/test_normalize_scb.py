"""Tests for the SCB (Sweden) workbook normalizer.

The real SCB workbook puts headers on row 11 and data from row 14, with the
Swedish and English name lists side by side and a column layout that changes
between year ranges. These tests build synthetic sheets with that shape so the
parsing rules are exercised without needing the 583 KB source file.
"""

import openpyxl
import pytest

from pipeline.normalize_scb import (
    _extract_year_from_sheet_name,
    _get_column_indices,
    _is_boys_sheet,
    _is_girls_sheet,
    _normalize_scb_sheet,
    normalize_scb_excel,
)


HEADER_ROW = 11
DATA_START_ROW = 14


def _build_sheet(wb, title, rows, num_cols=9, name_col=1, count_col=2):
    """Create a sheet shaped like SCB's: headers on row 11, data from row 14.

    `rows` is a list of (rank, name, count) written into the Swedish columns.
    """
    ws = wb.create_sheet(title)
    for c in range(1, num_cols + 1):
        ws.cell(row=HEADER_ROW, column=c, value=f"h{c}")
    for offset, (rank, name, count) in enumerate(rows):
        r = DATA_START_ROW + offset
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=name_col + 1, value=name)
        ws.cell(row=r, column=count_col + 1, value=count)
    return ws


class TestSheetNaming:
    def test_girls_sheet_recognised(self):
        assert _is_girls_sheet("Flickor 1998")
        assert not _is_boys_sheet("Flickor 1998")

    def test_boys_sheet_recognised(self):
        assert _is_boys_sheet("Pojkar 2021")
        assert not _is_girls_sheet("Pojkar 2021")

    def test_contents_sheet_is_neither(self):
        assert not _is_girls_sheet("Innehåll")
        assert not _is_boys_sheet("Innehåll")

    def test_year_extracted(self):
        assert _extract_year_from_sheet_name("Flickor 1998") == 1998
        assert _extract_year_from_sheet_name("Pojkar 2021") == 2021

    def test_year_missing_for_non_data_sheet(self):
        assert _extract_year_from_sheet_name("Innehåll") is None


class TestColumnLayout:
    """The workbook changes layout across year ranges; each must map correctly."""

    @pytest.mark.parametrize(
        "num_cols,expected_name,expected_count",
        [(9, 1, 2), (10, 1, 2), (11, 2, 3), (12, 2, 3), (13, 2, 3)],
    )
    def test_known_layouts(self, num_cols, expected_name, expected_count):
        name_col, count_col, _, _ = _get_column_indices(num_cols)
        assert (name_col, count_col) == (expected_name, expected_count)

    def test_unknown_layout_raises(self):
        with pytest.raises(ValueError):
            _get_column_indices(3)


class TestSheetNormalization:
    def test_extracts_canonical_records(self):
        wb = openpyxl.Workbook()
        ws = _build_sheet(wb, "Flickor 2021", [(1, "Alice", 706), (2, "Maja", 690)])

        records = _normalize_scb_sheet(ws)

        assert records == [
            {"name": "Alice", "country": "SE", "sex": "Girl", "year": 2021,
             "count": 706, "rank": 1},
            {"name": "Maja", "country": "SE", "sex": "Girl", "year": 2021,
             "count": 690, "rank": 2},
        ]

    def test_boys_sheet_maps_to_boy(self):
        wb = openpyxl.Workbook()
        ws = _build_sheet(wb, "Pojkar 2021", [(1, "Noah", 745)])

        assert _normalize_scb_sheet(ws)[0]["sex"] == "Boy"

    def test_tied_ranks_are_preserved(self):
        """SCB repeats a rank when names tie; both rows must survive."""
        wb = openpyxl.Workbook()
        ws = _build_sheet(wb, "Flickor 2021", [(10, "Ida", 749), (10, "Johanna", 749)])

        records = _normalize_scb_sheet(ws)

        assert [r["rank"] for r in records] == [10, 10]
        assert {r["name"] for r in records} == {"Ida", "Johanna"}

    def test_rows_without_a_name_are_skipped(self):
        wb = openpyxl.Workbook()
        ws = _build_sheet(wb, "Flickor 2021", [(1, "Alice", 706), (2, None, 690)])

        records = _normalize_scb_sheet(ws)

        assert len(records) == 1
        assert records[0]["name"] == "Alice"

    def test_non_numeric_rank_is_skipped(self):
        """Footnote rows sit below the data and must not become records."""
        wb = openpyxl.Workbook()
        ws = _build_sheet(wb, "Flickor 2021", [(1, "Alice", 706), ("Källa: SCB", "x", 1)])

        records = _normalize_scb_sheet(ws)

        assert len(records) == 1


class TestWorkbookNormalization:
    def test_combines_sheets_and_ignores_contents(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Innehåll").cell(row=1, column=1, value="contents")
        _build_sheet(wb, "Flickor 2020", [(1, "Alice", 700)])
        _build_sheet(wb, "Pojkar 2020", [(1, "Noah", 740)])
        path = tmp_path / "scb.xlsx"
        wb.save(path)

        df = normalize_scb_excel(path)

        assert len(df) == 2
        assert set(df.columns) >= {"name", "country", "sex", "year", "count", "rank"}
        assert set(df["sex"]) == {"Girl", "Boy"}
        assert set(df["country"]) == {"SE"}
