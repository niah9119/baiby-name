"""Tests for the ONS (England and Wales) workbook normalizer.

The real ONS workbook has headers on row 5, data from row 6, with one row per name
and pairs of rank/count columns for each year (2025 down to 1996). These tests
build synthetic sheets with that shape so the parsing rules are exercised without
needing the 12 MB source file.

Key behaviors tested:
- [x] values are skipped (suppressed data)
- Ranks are used directly (not computed)
- Sex is determined from sheet name (Table_1 = girls, Table_2 = boys)
"""

import openpyxl
import pytest

from pathlib import Path

from pipeline.normalize_ons import (
    _is_suppressed,
    _normalize_ons_sheet,
    normalize_ons_excel,
    GIRLS_SHEET_NAME,
    BOYS_SHEET_NAME,
    YEARS,
)

HEADER_ROW = 5
DATA_START_ROW = 6


def _build_sheet(wb, title, rows):
    """Create a sheet shaped like ONS's: headers on row 5, data from row 6.

    `rows` is a list of (name, [(year, rank, count), ...]) tuples.
    """
    ws = wb.create_sheet(title)
    # Header row (row 5, 1-indexed)
    ws.cell(row=HEADER_ROW, column=1, value="Name")
    for year_idx, year in enumerate(range(2025, 1995, -1)):
        rank_col = 2 + 2 * year_idx
        count_col = 3 + 2 * year_idx
        ws.cell(row=HEADER_ROW, column=rank_col, value=f"{year} Rank")
        ws.cell(row=HEADER_ROW, column=count_col, value=f"{year} Count")

    # Data rows
    for row_idx, (name, year_data) in enumerate(rows, start=DATA_START_ROW):
        ws.cell(row=row_idx, column=1, value=name)
        for year, rank, count in year_data:
            year_idx = 2025 - year
            rank_col = 2 + 2 * year_idx
            count_col = 3 + 2 * year_idx
            ws.cell(row=row_idx, column=rank_col, value=rank)
            ws.cell(row=row_idx, column=count_col, value=count)

    return ws


def _build_simple_sheet(wb, title, name_data_list):
    """Create a simpler sheet with just a few year columns.

    `name_data_list` is a list of (name, [(year, rank, count), ...]) tuples.
    """
    ws = wb.create_sheet(title)
    # Header row (row 5, 1-indexed)
    ws.cell(row=HEADER_ROW, column=1, value="Name")
    # Just do a few years for simplicity
    test_years = [2025, 2024, 2023]
    for year_idx, year in enumerate(test_years):
        rank_col = 2 + 2 * year_idx
        count_col = 3 + 2 * year_idx
        ws.cell(row=HEADER_ROW, column=rank_col, value=f"{year} Rank")
        ws.cell(row=HEADER_ROW, column=count_col, value=f"{year} Count")

    # Data rows
    for row_idx, (name, year_data) in enumerate(name_data_list, start=DATA_START_ROW):
        ws.cell(row=row_idx, column=1, value=name)
        for year, rank, count in year_data:
            year_idx = 2025 - year
            rank_col = 2 + 2 * year_idx
            count_col = 3 + 2 * year_idx
            ws.cell(row=row_idx, column=rank_col, value=rank)
            ws.cell(row=row_idx, column=count_col, value=count)

    return ws


class TestSuppression:
    def test_is_suppressed_detects_x(self):
        assert _is_suppressed("[x]") is True
        assert _is_suppressed("x") is False
        assert _is_suppressed(None) is False
        assert _is_suppressed(5) is False


class TestSheetNormalization:
    def test_extracts_canonical_records(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _build_simple_sheet(wb, GIRLS_SHEET_NAME, [
            ("Alice", [(2025, 10, 706), (2024, 11, 690)]),
        ])
        _build_simple_sheet(wb, BOYS_SHEET_NAME, [
            ("Noah", [(2025, 1, 745), (2024, 2, 720)]),
        ])

        path = tmp_path / "ons.xlsx"
        wb.save(path)
        df = normalize_ons_excel(path)

        assert len(df) == 4  # 2 names x 2 years each
        assert set(df["name"]) == {"Alice", "Noah"}
        assert set(df["sex"]) == {"Girl", "Boy"}
        assert set(df["country"]) == {"GB"}
        assert set(df["year"]) == {2024, 2025}

    def test_girls_sheet_maps_to_girl(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _build_simple_sheet(wb, GIRLS_SHEET_NAME, [("Alice", [(2025, 10, 706)])])

        path = tmp_path / "ons.xlsx"
        wb.save(path)
        df = normalize_ons_excel(path)
        assert df[df["name"] == "Alice"]["sex"].iloc[0] == "Girl"

    def test_boys_sheet_maps_to_boy(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _build_simple_sheet(wb, BOYS_SHEET_NAME, [("Noah", [(2025, 1, 745)])])

        path = tmp_path / "ons.xlsx"
        wb.save(path)
        df = normalize_ons_excel(path)
        assert df[df["name"] == "Noah"]["sex"].iloc[0] == "Boy"

    def test_suppressed_values_are_skipped(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        # Alice has [x] for 2025, actual data for 2024
        _build_simple_sheet(wb, GIRLS_SHEET_NAME, [
            ("Alice", [(2025, "[x]", "[x]"), (2024, 11, 690)]),
        ])

        path = tmp_path / "ons.xlsx"
        wb.save(path)
        df = normalize_ons_excel(path)
        alice_rows = df[df["name"] == "Alice"]
        assert len(alice_rows) == 1
        assert alice_rows["year"].iloc[0] == 2024
        assert alice_rows["count"].iloc[0] == 690

    def test_rank_is_published_not_computed(self, tmp_path):
        """The source provides ranks, we don't recompute them."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _build_simple_sheet(wb, GIRLS_SHEET_NAME, [
            ("Alice", [(2025, 100, 50)]),  # rank 100
            ("Bob", [(2025, 1, 100)]),  # rank 1 (same count, different rank)
        ])

        path = tmp_path / "ons.xlsx"
        wb.save(path)
        df = normalize_ons_excel(path)
        ranks = dict(zip(df["name"], df["rank"]))
        # Ranks should be exactly as in the source
        assert ranks["Alice"] == 100
        assert ranks["Bob"] == 1

    def test_multiple_years_per_name(self, tmp_path):
        """Names can appear in multiple years with different ranks."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _build_simple_sheet(wb, GIRLS_SHEET_NAME, [
            ("Emma", [(2025, 1, 379), (2024, 2, 366), (2023, 1, 401)]),
        ])

        path = tmp_path / "ons.xlsx"
        wb.save(path)
        df = normalize_ons_excel(path)
        emma_rows = df[df["name"] == "Emma"]
        assert len(emma_rows) == 3
        years = dict(zip(emma_rows["year"], emma_rows["rank"]))
        assert years[2025] == 1
        assert years[2024] == 2
        assert years[2023] == 1

    def test_rows_without_name_are_skipped(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _build_simple_sheet(wb, GIRLS_SHEET_NAME, [
            ("Alice", [(2025, 10, 706)]),
            (None, [(2025, 11, 690)]),  # No name - should be skipped
        ])

        path = tmp_path / "ons.xlsx"
        wb.save(path)
        df = normalize_ons_excel(path)
        assert len(df) == 1
        assert df["name"].iloc[0] == "Alice"


class TestWorkbookNormalization:
    def test_combines_sheets_and_ignores_other_sheets(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        # Non-data sheets should be ignored
        wb.create_sheet("Cover_sheet").cell(row=1, column=1, value="info")
        wb.create_sheet("Contents").cell(row=1, column=1, value="contents")
        wb.create_sheet("Notes").cell(row=1, column=1, value="notes")
        _build_simple_sheet(wb, GIRLS_SHEET_NAME, [("Alice", [(2025, 100, 700)])])
        _build_simple_sheet(wb, BOYS_SHEET_NAME, [("Noah", [(2025, 1, 740)])])

        path = tmp_path / "ons.xlsx"
        wb.save(path)

        df = normalize_ons_excel(path)

        assert len(df) == 2
        assert set(df.columns) >= {"name", "country", "sex", "year", "count", "rank"}
        assert set(df["sex"]) == {"Girl", "Boy"}
        assert set(df["country"]) == {"GB"}

    def test_all_columns_present(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _build_simple_sheet(wb, GIRLS_SHEET_NAME, [("Alice", [(2025, 1, 706)])])
        path = tmp_path / "ons.xlsx"
        wb.save(path)

        df = normalize_ons_excel(path)

        expected_cols = ["name", "country", "sex", "year", "count", "rank"]
        assert list(df.columns) == expected_cols

    def test_suppressed_count_is_skipped(self, tmp_path):
        """When count is [x], the entire record is skipped."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _build_simple_sheet(wb, GIRLS_SHEET_NAME, [
            ("Alice", [(2025, 10, 706), (2024, "[x]", "[x]"), (2023, 11, 690)]),
        ])

        path = tmp_path / "ons.xlsx"
        wb.save(path)
        df = normalize_ons_excel(path)

        alice_rows = df[df["name"] == "Alice"]
        years = set(alice_rows["year"])
        assert 2024 not in years  # Suppressed year should be skipped
        assert 2025 in years
        assert 2023 in years
