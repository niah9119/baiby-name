"""Tests for canonical sex vocabulary across all importers.

The canonical sex vocabulary is 'Boy' and 'Girl'. All importers must emit
exactly these two values, not any alternatives (e.g., 'M'/'F', 'boy'/'girl').
"""

import csv
import tempfile
from pathlib import Path

import pandas as pd
import pytest

from pipeline.normalize import normalize_ssa_file
from pipeline.normalize_dst import normalize_dst_file, DENMARK_COUNTRY_CODE
from pipeline.normalize_ons import normalize_ons_excel, GIRLS_SHEET_NAME, BOYS_SHEET_NAME
from pipeline.normalize_scb import normalize_scb_excel
from pipeline.normalize_ssb import normalize_ssb_dataset
import openpyxl
import json


CANONICAL_SEX_VALUES = {"Boy", "Girl"}


class TestCanonicalSexValues:
    """Test that all importers produce only the canonical sex values."""

    def test_ssa_normalizes_m_to_boy(self):
        """SSA 'M' should be normalized to 'Boy'."""
        with tempfile.TemporaryDirectory() as tmpdir:
            ssa_file = Path(tmpdir) / "yob2023.txt"
            ssa_file.write_text("Bob,M,1500\n")

            df = normalize_ssa_file(ssa_file)

            assert df["sex"].unique().tolist() == ["Boy"]

    def test_ssa_normalizes_f_to_girl(self):
        """SSA 'F' should be normalized to 'Girl'."""
        with tempfile.TemporaryDirectory() as tmpdir:
            ssa_file = Path(tmpdir) / "yob2023.txt"
            ssa_file.write_text("Alice,F,1000\n")

            df = normalize_ssa_file(ssa_file)

            assert df["sex"].unique().tolist() == ["Girl"]

    def test_ssa_mixed_normalized(self):
        """SSA mixed M/F should be normalized to Boy/Girl."""
        with tempfile.TemporaryDirectory() as tmpdir:
            ssa_file = Path(tmpdir) / "yob2023.txt"
            ssa_file.write_text("Alice,F,1000\nBob,M,1500\nCharlie,F,500\n")

            df = normalize_ssa_file(ssa_file)

            assert set(df["sex"].unique()) == CANONICAL_SEX_VALUES

    def test_ons_sex_values_are_canonical(self, tmp_path):
        """ONS should produce 'Boy' and 'Girl'."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(GIRLS_SHEET_NAME)
        # Header on row 5
        ws.cell(row=5, column=1, value="Name")
        ws.cell(row=5, column=2, value="2025 Rank")
        ws.cell(row=5, column=3, value="2025 Count")
        # Data on row 6
        ws.cell(row=6, column=1, value="Alice")
        ws.cell(row=6, column=2, value=1)
        ws.cell(row=6, column=3, value=100)

        ws = wb.create_sheet(BOYS_SHEET_NAME)
        ws.cell(row=5, column=1, value="Name")
        ws.cell(row=5, column=2, value="2025 Rank")
        ws.cell(row=5, column=3, value="2025 Count")
        ws.cell(row=6, column=1, value="Bob")
        ws.cell(row=6, column=2, value=1)
        ws.cell(row=6, column=3, value=100)

        path = tmp_path / "ons.xlsx"
        wb.save(path)

        df = normalize_ons_excel(path)

        assert set(df["sex"].unique()) == CANONICAL_SEX_VALUES

    def test_scb_sex_values_are_canonical(self, tmp_path):
        """SCB should produce 'Boy' and 'Girl'."""
        # Use the existing test helper pattern from test_normalize_scb.py
        # to create a valid SCB Excel file
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Flickor 2020")
        # Headers on row 11, data starts on row 14
        # Layout for 13 columns: Rank, Extra, Name, Count, Per1000, None, Name(EN), Count(EN), Per1000(EN), Rank(EN), Extra, Extra, Extra
        ws.cell(row=11, column=1, value="Rank")
        ws.cell(row=11, column=2, value="Extra")
        ws.cell(row=11, column=3, value="Namn")  # Swedish name column
        ws.cell(row=11, column=4, value="Antal")  # Swedish count column
        for c in range(5, 14):
            ws.cell(row=11, column=c, value=f"Col{c}")

        ws.cell(row=14, column=1, value=1)  # Rank
        ws.cell(row=14, column=3, value="Alice")  # Name (column 3 for 13-col layout)
        ws.cell(row=14, column=4, value=100)  # Count

        ws = wb.create_sheet("Pojkar 2020")
        for c in range(1, 14):
            ws.cell(row=11, column=c, value=f"Col{c}")
        ws.cell(row=14, column=1, value=1)
        ws.cell(row=14, column=3, value="Bob")
        ws.cell(row=14, column=4, value=100)

        path = tmp_path / "scb.xlsx"
        wb.save(path)

        df = normalize_scb_excel(path)

        assert set(df["sex"].unique()) == CANONICAL_SEX_VALUES

    def test_ssb_sex_values_are_canonical(self, tmp_path):
        """SSB should produce 'Boy' and 'Girl'."""
        payload = {
            "id": ["Fornavn", "ContentsCode", "Tid"],
            "size": [2, 1, 1],
            "value": [100, 150],
            "dimension": {
                "Fornavn": {
                    "category": {
                        "index": {"1EMMA": 0, "2JAKOB": 1},
                        "label": {"1EMMA": "Emma", "2JAKOB": "Jakob"},
                    }
                },
                "ContentsCode": {"category": {"index": {"Personer": 0}}},
                "Tid": {"category": {"index": {"2024": 0}}},
            },
        }
        path = tmp_path / "ssb.json"
        path.write_text(json.dumps(payload), encoding="utf-8")

        df = normalize_ssb_dataset(path)

        assert set(df["sex"].unique()) == CANONICAL_SEX_VALUES

    def test_dst_sex_values_are_canonical(self, tmp_path):
        """DST should produce 'Boy' and 'Girl'."""
        html = """
        <table class="table">
            <caption class="names__headerName">Pigenavne</caption>
            <thead><tr><th>Nr</th><th>Navn</th><th>Antal</th><th>Pr. 1 000</th></tr></thead>
            <tbody>
                <tr><td>1</td><td>Emma</td><td>445</td><td>16</td></tr>
            </tbody>
        </table>
        <table class="table">
            <caption class="names__headerName">Drengenavne</caption>
            <thead><tr><th>Nr</th><th>Navn</th><th>Antal</th><th>Pr. 1 000</th></tr></thead>
            <tbody>
                <tr><td>1</td><td>Oscar</td><td>450</td><td>16</td></tr>
            </tbody>
        </table>
        """
        file_path = tmp_path / "dst-2024.html"
        file_path.write_text(html, encoding="utf-8")

        df = normalize_dst_file(file_path)

        assert set(df["sex"].unique()) == CANONICAL_SEX_VALUES
