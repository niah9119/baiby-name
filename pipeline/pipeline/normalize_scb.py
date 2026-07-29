"""Normalize SCB (Statistics Sweden) baby names data to canonical CSV format."""

import re
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

from .config import CANONICAL_CSV_PATH, SCB_DATA_DIR, SWEDEN_COUNTRY_CODE

# Sheet name patterns for girls and boys
GIRLS_SHEET_PATTERN = re.compile(r"^Flickor (\d{4})$")
BOYS_SHEET_PATTERN = re.compile(r"^Pojkar (\d{4})$")

# Column structure varies by year:
# 1998-2000 (9 cols): Rank, Name, Count, Per1000, None, Name(EN), Count(EN), Per1000(EN), Rank(EN)
#   Swedish: Name=col1, Count=col2
#   English: Name=col5, Count=col6
# 2001-2017 (11 cols): Rank, Extra, Name, Count, Per1000, None, Name(EN), Count(EN), Per1000(EN), Rank(EN), Extra
#   Swedish: Name=col2, Count=col3
#   English: Name=col6, Count=col7
# 2019+ (13 cols): Similar to 11 cols with extra columns
#   Swedish: Name=col2, Count=col3
#   English: Name=col6, Count=col8

# Column indices for Swedish data
SWEDISH_NAME_COL_9 = 1  # For 9-column structure (1998-2000)
SWEDISH_COUNT_COL_9 = 2

SWEDISH_NAME_COL_11 = 2  # For 11-column structure (2001-2017)
SWEDISH_COUNT_COL_11 = 3

SWEDISH_NAME_COL_13 = 2  # For 13-column structure (2019+)
SWEDISH_COUNT_COL_13 = 3


def _is_scb_excel_file(file_path: Path) -> bool:
    """Check if a file is the expected SCB Excel file."""
    return file_path.suffix.lower() in (".xlsx", ".xls")


def _extract_year_from_sheet_name(sheet_name: str) -> Optional[int]:
    """Extract year from sheet name like 'Flickor 1998' or 'Pojkar 2021'."""
    match = re.search(r"(\d{4})$", sheet_name)
    if match:
        return int(match.group(1))
    return None


def _is_girls_sheet(sheet_name: str) -> bool:
    """Check if sheet is for girls (Flickor)."""
    return bool(GIRLS_SHEET_PATTERN.match(sheet_name))


def _is_boys_sheet(sheet_name: str) -> bool:
    """Check if sheet is for boys (Pojkar)."""
    return bool(BOYS_SHEET_PATTERN.match(sheet_name))


def _get_column_indices(num_cols: int) -> tuple[int, int, int, int]:
    """
    Get column indices based on number of columns in the sheet.

    Returns:
        Tuple of (swedish_name_col, swedish_count_col, english_name_col, english_count_col)
    """
    if num_cols == 9:
        return (SWEDISH_NAME_COL_9, SWEDISH_COUNT_COL_9,
                SWEDISH_NAME_COL_9 + 4, SWEDISH_COUNT_COL_9 + 4)
    elif num_cols == 10:
        # 2000 Pojkar has 10 columns - treat like 9
        return (SWEDISH_NAME_COL_9, SWEDISH_COUNT_COL_9,
                SWEDISH_NAME_COL_9 + 4, SWEDISH_COUNT_COL_9 + 4)
    elif num_cols == 11:
        return (SWEDISH_NAME_COL_11, SWEDISH_COUNT_COL_11,
                SWEDISH_NAME_COL_11 + 4, SWEDISH_COUNT_COL_11 + 4)
    elif num_cols == 12:
        # 2018 Pojkar - slightly different layout
        # Swedish: Name=col2, Count=col3
        # English: Name=col6, Count=col7
        return (2, 3, 6, 7)
    elif num_cols >= 13:
        return (SWEDISH_NAME_COL_13, SWEDISH_COUNT_COL_13,
                SWEDISH_NAME_COL_13 + 4, SWEDISH_COUNT_COL_13 + 4)
    else:
        raise ValueError(f"Unknown column structure with {num_cols} columns")


def _normalize_scb_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """
    Normalize a single SCB Excel sheet to canonical format.

    The sheet contains Swedish and English name statistics in a dual-column format.
    We extract the Swedish data which contains the rank and count.

    Data starts at row 14 (1-indexed), with headers at row 11.

    Args:
        ws: Openpyxl worksheet object

    Returns:
        List of dictionaries in canonical format
    """
    records = []

    # Detect column structure by examining headers or total columns
    num_cols = len(list(ws[11]))
    swedish_name_col, swedish_count_col, _, _ = _get_column_indices(num_cols)

    # Find the data start row (first row with a numeric rank in column 0)
    # Row 11 has headers, data starts at row 14
    start_row = 14

    for row_idx in range(start_row, ws.max_row + 1):
        row = ws[row_idx]

        # Check if we've hit the end of data
        rank_cell = row[0]
        if rank_cell.value is None:
            # Empty cell - likely end of data
            continue

        try:
            rank = int(rank_cell.value)
        except (ValueError, TypeError):
            # Not a valid rank, skip
            continue

        # Extract data from Swedish columns
        name = row[swedish_name_col].value
        count = row[swedish_count_col].value

        # Skip if name or count is missing
        if not name or count is None:
            continue

        # Extract year from sheet name
        year = _extract_year_from_sheet_name(ws.title)

        # Determine sex from sheet type
        if _is_girls_sheet(ws.title):
            sex = "Girl"
        elif _is_boys_sheet(ws.title):
            sex = "Boy"
        else:
            continue  # Skip unknown sheet types

        records.append({
            "name": str(name).strip(),
            "country": SWEDEN_COUNTRY_CODE,
            "sex": sex,
            "year": year,
            "count": int(count),
            "rank": rank,
        })

    return records


def normalize_scb_excel(file_path: Path) -> pd.DataFrame:
    """
    Normalize SCB Excel file to canonical CSV format.

    Args:
        file_path: Path to the SCB Excel file (scb-nyfodda-1998-2021.xlsx)

    Returns:
        DataFrame in canonical format
    """
    wb = openpyxl.load_workbook(file_path)

    all_records = []

    # Process each sheet
    for sheet_name in wb.sheetnames:
        # Skip the Contents sheet and any non-data sheets
        if sheet_name == "Innehåll":
            continue

        # Check if this is a data sheet for girls or boys
        if not (_is_girls_sheet(sheet_name) or _is_boys_sheet(sheet_name)):
            continue

        ws = wb[sheet_name]
        records = _normalize_scb_sheet(ws)
        all_records.extend(records)

    # Create DataFrame
    df = pd.DataFrame(all_records)

    if df.empty:
        raise ValueError("No data found in Excel file")

    # Ensure correct column order
    df = df[["name", "country", "sex", "year", "count", "rank"]]

    # Sort by country, sex, year, rank for consistent output
    df = df.sort_values(
        ["country", "sex", "year", "rank"], ascending=[True, True, True, True]
    )

    return df


def normalize_all_files(
    input_dir: Optional[Path] = None, output_path: Optional[Path] = None
) -> pd.DataFrame:
    """
    Normalize all SCB Excel files in the input directory.

    Args:
        input_dir: Directory containing SCB data files
        output_path: Path to write the canonical CSV

    Returns:
        DataFrame with all normalized data
    """
    input_dir = input_dir or SCB_DATA_DIR
    output_path = output_path or CANONICAL_CSV_PATH

    # Find SCB Excel files
    excel_files = sorted(input_dir.glob("*.xlsx")) + sorted(input_dir.glob("*.xls"))

    if not excel_files:
        raise ValueError(f"No SCB Excel files found in {input_dir}")

    # Process all Excel files and combine
    all_dfs = []
    for file_path in excel_files:
        print(f"Processing: {file_path.name}")
        df = normalize_scb_excel(file_path)
        print(f"  Found {len(df)} records")
        all_dfs.append(df)

    # Combine all dataframes
    combined_df = pd.concat(all_dfs, ignore_index=True)

    # Write to CSV
    output_path.parent.mkdir(parents=True, exist_ok=True)
    combined_df.to_csv(output_path, index=False)

    print(f"Canonical CSV written to: {output_path}")
    print(f"Total records: {len(combined_df)}")

    return combined_df


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Normalize SCB Excel data to canonical CSV")
    parser.add_argument("--input-dir", type=str, help="Directory with SCB Excel files")
    parser.add_argument("--input-file", type=str, help="Specific SCB Excel file to process")
    parser.add_argument("--output", type=str, help="Output CSV path")

    args = parser.parse_args()

    input_dir = Path(args.input_dir) if args.input_dir else None
    output_path = Path(args.output) if args.output else None

    if args.input_file:
        # Process a specific file
        file_path = Path(args.input_file)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        df = normalize_scb_excel(file_path)
        if output_path:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(output_path, index=False)
            print(f"Output written to: {output_path}")
    else:
        df = normalize_all_files(input_dir, output_path)

    print(f"\nSummary:")
    print(f"  Total names: {df['name'].nunique()}")
    print(f"  Country coverage: {df['country'].unique()}")
    print(f"  Sex categories: {df['sex'].unique()}")
    print(f"  Year range: {df['year'].min()} - {df['year'].max()}")
