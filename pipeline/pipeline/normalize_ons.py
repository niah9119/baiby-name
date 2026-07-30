"""Normalize ONS (England and Wales) baby names data to canonical CSV format.

The source Excel workbook has 5 sheets:
- Cover_sheet: Informational
- Contents: Table of contents
- Notes: footnotes
- Table_1: Names for baby girls (24,007 rows, years 1996-2025)
- Table_2: Names for baby boys (18,087 rows, years 1996-2025)

The table layout is "wide" format with one row per name and pairs of rank/count columns
for each year. We must unpivot this into canonical long format (one row per name/year/sex).

Key details:
- Header is on row 5, data starts on row 6
- [x] means suppressed (below disclosure threshold) - skip these, don't convert to 0
- Ranks are published, so use them directly (don't compute)
"""

from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

from .config import CANONICAL_CSV_PATH, GB_ENGLAND_WALES_COUNTRY_CODE, ONS_DATA_DIR

# Sheet names for girls and boys
GIRLS_SHEET_NAME = "Table_1"
BOYS_SHEET_NAME = "Table_2"

# Years covered by the workbook (descending from 2025 to 1996)
YEARS = list(range(2025, 1995, -1))


def _is_suppressed(value) -> bool:
    """Check if a cell value indicates suppressed data ([x])."""
    return value == "[x]"


def _normalize_ons_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """
    Normalize a single ONS Excel sheet to canonical format.

    Args:
        ws: Openpyxl worksheet object (Table_1 or Table_2)

    Returns:
        List of dictionaries in canonical format
    """
    records = []

    # Sheet titles indicate the sex
    if ws.title == GIRLS_SHEET_NAME:
        sex = "Girl"
    elif ws.title == BOYS_SHEET_NAME:
        sex = "Boy"
    else:
        return []

    # Data starts on row 6 (1-indexed)
    # Column 1 = Name
    # Columns 2,3 = 2025 Rank, 2025 Count
    # Columns 4,5 = 2024 Rank, 2024 Count
    # ... and so on down to 1996

    for row_idx in range(6, ws.max_row + 1):
        # Get the name (column 1, 1-indexed)
        name_cell = ws.cell(row=row_idx, column=1)
        name = name_cell.value

        # Skip rows without a name or header
        if not name or name == "Name":
            continue

        name = str(name).strip()

        # Process each year's rank/count pair
        for year_idx, year in enumerate(YEARS):
            # Column indices (1-indexed):
            # Name = 1
            # 2025 Rank = 2, 2025 Count = 3
            # 2024 Rank = 4, 2024 Count = 5
            # ...
            # year Rank = 2 + 2*year_idx, year Count = 3 + 2*year_idx
            rank_col = 2 + 2 * year_idx
            count_col = 3 + 2 * year_idx

            rank_cell = ws.cell(row=row_idx, column=rank_col)
            count_cell = ws.cell(row=row_idx, column=count_col)

            rank = rank_cell.value
            count = count_cell.value

            # Skip suppressed values
            if _is_suppressed(rank) or _is_suppressed(count):
                continue

            # Skip if count is None or not a number
            if count is None:
                continue
            try:
                count = int(count)
            except (ValueError, TypeError):
                continue

            # Get rank - if suppressed, skip the whole record
            if _is_suppressed(rank):
                continue
            try:
                rank = int(rank)
            except (ValueError, TypeError):
                continue

            records.append({
                "name": name,
                "country": GB_ENGLAND_WALES_COUNTRY_CODE,
                "sex": sex,
                "year": year,
                "count": count,
                "rank": rank,
            })

    return records


def normalize_ons_excel(file_path: Path) -> pd.DataFrame:
    """
    Normalize ONS Excel file to canonical CSV format.

    Args:
        file_path: Path to the ONS Excel file (ons-babynames-1996-2025.xlsx)

    Returns:
        DataFrame in canonical format
    """
    wb = openpyxl.load_workbook(file_path)

    all_records = []

    # Process Table_1 (girls) and Table_2 (boys)
    for sheet_name in [GIRLS_SHEET_NAME, BOYS_SHEET_NAME]:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        records = _normalize_ons_sheet(ws)
        all_records.extend(records)

    # Create DataFrame
    df = pd.DataFrame(all_records)

    if df.empty:
        raise ValueError("No data found in Excel file")

    # Ensure correct column order
    df = df[["name", "country", "sex", "year", "count", "rank"]]

    # Sort by country, sex, year, rank for consistent output
    df = df.sort_values(
        ["country", "sex", "year", "rank"], ascending=[True, True, True, True]
    )

    return df


def normalize_all_files(
    input_dir: Optional[Path] = None, output_path: Optional[Path] = None
) -> pd.DataFrame:
    """
    Normalize all ONS Excel files in the input directory.

    Args:
        input_dir: Directory containing ONS data files
        output_path: Path to write the canonical CSV

    Returns:
        DataFrame with all normalized data
    """
    input_dir = input_dir or ONS_DATA_DIR
    output_path = output_path or CANONICAL_CSV_PATH

    # Find ONS Excel files
    excel_files = sorted(input_dir.glob("*.xlsx")) + sorted(input_dir.glob("*.xls"))

    if not excel_files:
        raise ValueError(f"No ONS Excel files found in {input_dir}")

    # Process all Excel files and combine
    all_dfs = []
    for file_path in excel_files:
        print(f"Processing: {file_path.name}")
        df = normalize_ons_excel(file_path)
        print(f"  Found {len(df)} records")
        all_dfs.append(df)

    # Combine all dataframes
    combined_df = pd.concat(all_dfs, ignore_index=True)

    # Write to CSV
    output_path.parent.mkdir(parents=True, exist_ok=True)
    combined_df.to_csv(output_path, index=False)

    print(f"Canonical CSV written to: {output_path}")
    print(f"Total records: {len(combined_df)}")

    return combined_df


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Normalize ONS Excel data to canonical CSV"
    )
    parser.add_argument("--input-dir", type=str, help="Directory with ONS Excel files")
    parser.add_argument("--input-file", type=str, help="Specific ONS Excel file to process")
    parser.add_argument("--output", type=str, help="Output CSV path")

    args = parser.parse_args()

    input_dir = Path(args.input_dir) if args.input_dir else None
    output_path = Path(args.output) if args.output else None

    if args.input_file:
        # Process a specific file
        file_path = Path(args.input_file)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        df = normalize_ons_excel(file_path)
        if output_path:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(output_path, index=False)
            print(f"Output written to: {output_path}")
    else:
        df = normalize_all_files(input_dir, output_path)

    print(f"\nSummary:")
    print(f"  Total names: {df['name'].nunique()}")
    print(f"  Country coverage: {df['country'].unique()}")
    print(f"  Sex categories: {df['sex'].unique()}")
    print(f"  Year range: {df['year'].min()} - {df['year'].max()}")
